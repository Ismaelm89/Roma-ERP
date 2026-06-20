"""Import items (and variants) from an Excel file.

Expected columns (first row is the header — order doesn't matter):
    code           — model code, e.g. "101"               (required)
    name_ar        — Arabic name                          (required)
    name_en        — English name                         (optional)
    category       — e.g. "pajama set"                    (optional)
    fabric         — free text, e.g. "قطن 100%", "بوليستر" (optional, default blank)
    description    — free text                            (optional)
    sizes          — comma-separated, e.g. "S,M,L,XL" or "10,12,14" (optional, default "S,M,L,XL,XXL")
    active         — TRUE/FALSE/1/0                       (optional, default TRUE)

Usage:
    python manage.py import_items <path-to-xlsx> [--dry-run]
"""
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from inventory.models import Item, ItemVariant


DEFAULT_SIZES = ['S', 'M', 'L', 'XL', 'XXL']


def _truthy(v):
    if v is None:
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ('1', 'true', 'yes', 'y', 't', 'نعم', 'حق')


def _str(v):
    return '' if v is None else str(v).strip()


class Command(BaseCommand):
    help = 'Import items + variants from an Excel (.xlsx) file.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Path to the .xlsx file')
        parser.add_argument('--dry-run', action='store_true',
                             help='Validate only — do not write to DB')

    def handle(self, *args, **opts):
        path = Path(opts['path'])
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('Empty workbook.')

        header = [(_str(c) or '').lower() for c in rows[0]]
        if 'code' not in header or 'name_ar' not in header:
            raise CommandError(
                'Required columns missing. Need at least: code, name_ar. '
                f'Found header: {header}'
            )
        col = {name: header.index(name) for name in header}

        created_items = updated_items = created_variants = skipped_sizes = 0
        errors = []
        # كتالوج المقاسات المعتمد — أي مقاس بره الكتالوج بيتتخطّى (حماية من بيانات غلط).
        from manufacturing.models import Size
        valid_sizes = set(Size.objects.values_list('code', flat=True))

        with transaction.atomic():
            for row_idx, row in enumerate(rows[1:], start=2):
                if all(c in (None, '') for c in row):
                    continue
                code = _str(row[col['code']])
                name_ar = _str(row[col['name_ar']])
                if not code or not name_ar:
                    errors.append(f'Row {row_idx}: missing code or name_ar')
                    continue

                defaults = {
                    'name_ar': name_ar,
                    'name_en': _str(row[col['name_en']]) if 'name_en' in col else '',
                    'category': _str(row[col['category']]) if 'category' in col else '',
                    'fabric': _str(row[col['fabric']]) if 'fabric' in col else '',
                    'description': _str(row[col['description']]) if 'description' in col else '',
                    'active': _truthy(row[col['active']]) if 'active' in col else True,
                }

                item, was_created = Item.objects.update_or_create(
                    code=code, defaults=defaults,
                )
                if was_created:
                    created_items += 1
                else:
                    updated_items += 1

                sizes_raw = _str(row[col['sizes']]) if 'sizes' in col else ''
                sizes = [s.strip().upper() for s in sizes_raw.split(',') if s.strip()] or DEFAULT_SIZES

                for size in sizes:
                    # تحقّق إن المقاس موجود في كتالوج المقاسات — يمنع إنشاء مقاس بالغلط
                    # من قيمة غلط في عمود المقاسات (زي توكن من اسم المنتج).
                    if valid_sizes and size not in valid_sizes:
                        self.stdout.write(self.style.WARNING(
                            f'  تخطّي مقاس غير معروف «{size}» للصنف {code} '
                            f'(مش في كتالوج المقاسات).'))
                        skipped_sizes += 1
                        continue
                    _, was_created = ItemVariant.objects.get_or_create(
                        item=item, size=size,
                    )
                    if was_created:
                        created_variants += 1

            if opts['dry_run']:
                self.stdout.write(self.style.WARNING('DRY RUN — rolling back.'))
                transaction.set_rollback(True)

        for e in errors:
            self.stdout.write(self.style.ERROR(e))
        self.stdout.write(self.style.SUCCESS(
            f'Items: {created_items} created, {updated_items} updated. '
            f'Variants created: {created_variants}. Errors: {len(errors)}.'
        ))
