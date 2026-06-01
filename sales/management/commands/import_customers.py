"""Import customers from an Excel file.

Expected columns (first row is the header — order doesn't matter):
    code                 — customer code                 (required)
    name_ar              — Arabic name                   (required)
    name_en              — English name                  (optional)
    tax_id               — tax ID                        (optional)
    commercial_register  — CR number                     (optional)
    address              — full address                  (optional)
    governorate          — governorate                   (optional)
    phone                — phone                         (optional)
    email                — email                         (optional)
    payment_terms_days   — int                           (optional, default 0)
    credit_limit         — decimal                       (optional, default 0)
    opening_balance      — decimal (positive = owed to us)(optional, default 0)
    active               — TRUE/FALSE                    (optional, default TRUE)

Usage:
    python manage.py import_customers <path-to-xlsx> [--dry-run]
"""
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from sales.models import Customer


def _str(v):
    return '' if v is None else str(v).strip()


def _decimal(v, default=Decimal('0')):
    if v in (None, ''):
        return default
    try:
        return Decimal(str(v))
    except InvalidOperation:
        return default


def _int(v, default=0):
    if v in (None, ''):
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def _truthy(v):
    if v is None:
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ('1', 'true', 'yes', 'y', 't', 'نعم', 'حق')


class Command(BaseCommand):
    help = 'Import customers from an Excel (.xlsx) file.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Path to the .xlsx file')
        parser.add_argument('--dry-run', action='store_true',
                             help='Validate only — do not write to DB')

    def handle(self, *args, **opts):
        path = Path(opts['path'])
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError('Empty workbook.')

        header = [(_str(c) or '').lower() for c in rows[0]]
        if 'code' not in header or 'name_ar' not in header:
            raise CommandError(
                'Required columns missing. Need at least: code, name_ar. '
                f'Found header: {header}'
            )
        col = {name: header.index(name) for name in header}

        def cell(row, name, default=None):
            return row[col[name]] if name in col and col[name] < len(row) else default

        created = updated = 0
        errors = []

        with transaction.atomic():
            for row_idx, row in enumerate(rows[1:], start=2):
                if all(c in (None, '') for c in row):
                    continue
                code = _str(cell(row, 'code'))
                name_ar = _str(cell(row, 'name_ar'))
                if not code or not name_ar:
                    errors.append(f'Row {row_idx}: missing code or name_ar')
                    continue

                defaults = {
                    'name_ar': name_ar,
                    'name_en': _str(cell(row, 'name_en')),
                    'tax_id': _str(cell(row, 'tax_id')),
                    'commercial_register': _str(cell(row, 'commercial_register')),
                    'address': _str(cell(row, 'address')),
                    'governorate': _str(cell(row, 'governorate')),
                    'phone': _str(cell(row, 'phone')),
                    'email': _str(cell(row, 'email')),
                    'payment_terms_days': _int(cell(row, 'payment_terms_days'), 0),
                    'credit_limit': _decimal(cell(row, 'credit_limit')),
                    'opening_balance': _decimal(cell(row, 'opening_balance')),
                    'active': _truthy(cell(row, 'active')),
                }

                _, was_created = Customer.objects.update_or_create(
                    code=code, defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            if opts['dry_run']:
                self.stdout.write(self.style.WARNING('DRY RUN — rolling back.'))
                transaction.set_rollback(True)

        for e in errors:
            self.stdout.write(self.style.ERROR(e))
        self.stdout.write(self.style.SUCCESS(
            f'Customers: {created} created, {updated} updated. Errors: {len(errors)}.'
        ))
